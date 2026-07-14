import openpyxl
from openpyxl.styles import Protection
import shutil
import os

def protect_template():
    src = r"d:\Resume Projects\Time Table Data.xlsx"
    dest_dir = r"d:\Resume Projects\timetable-engine\frontend"
    os.makedirs(dest_dir, exist_ok=True)
    dest = os.path.join(dest_dir, "Time_Table_Template.xlsx")
    
    # Load workbook
    wb = openpyxl.load_workbook(src)
    
    # Apply lock on modifications of structure
    from openpyxl.workbook.protection import WorkbookProtection
    wb.security = WorkbookProtection()
    wb.security.workbookPassword = "optischedule"
    wb.security.lockStructure = True
    
    # Lock/Unlock specific parts of sheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.protection.sheet = True
        ws.protection.password = "optischedule"
        
        # In Teachers Data, unlock columns A to I (rows 2 to 100) for entry
        if sheet_name == "Teachers Data":
            for r in range(2, 150):
                for c in range(1, 10):
                    ws.cell(row=r, column=c).protection = Protection(locked=False)
                    
        # In Lecture Count, unlock columns A to C (rows 2 to 100)
        elif sheet_name == "Lecture Count":
            for r in range(2, 100):
                for c in range(1, 4):
                    ws.cell(row=r, column=c).protection = Protection(locked=False)
                    
        # In Max Lecture Count for Teachers, unlock columns B, C, D, F, G, H (rows 2 to 30)
        elif sheet_name == "Max Lecture Count for Teachers":
            for r in range(2, 30):
                for c in [2, 3, 4, 6, 7, 8]:
                    ws.cell(row=r, column=c).protection = Protection(locked=False)
                    
        # In Sample Class Time Table, unlock everything so they can change times if wanted
        elif sheet_name == "Sample Class Time Table":
            for r in range(1, 30):
                for c in range(1, 8):
                    ws.cell(row=r, column=c).protection = Protection(locked=False)
                    
    wb.save(dest)
    print(f"Protected template created successfully at: {dest}")

if __name__ == "__main__":
    protect_template()
