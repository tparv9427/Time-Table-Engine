import openpyxl
import re
from io import BytesIO
from typing import Dict, List, Any, Tuple
import structlog

logger = structlog.get_logger()

def clean_str(val: Any) -> str:
    if val is None:
        return ""
    return str(val).strip()

def clean_int(val: Any) -> int:
    if val is None:
        return 0
    try:
        # Handle decimal string parsing, like "6.0"
        return int(float(str(val).strip()))
    except (ValueError, TypeError):
        return 0

class IngestionError(Exception):
    def __init__(self, message: str, errors: List[str] = None):
        super().__init__(message)
        self.errors = errors or [message]

def parse_teacher_constraint(constraint_str: str, slots_count: int, days_list: List[str]) -> Dict[str, Any]:
    """
    Parses natural language constraint descriptions from the user template.
    Defaults to full availability if constraint string is empty.
    """
    allowed_days = days_list.copy()
    allowed_slots = list(range(1, slots_count + 1))
    max_per_day = slots_count

    if not constraint_str:
        return {
            "allowed_days": allowed_days,
            "allowed_slots": allowed_slots,
            "max_per_day": max_per_day
        }

    c_lower = constraint_str.lower()
    
    # 1. Parse Slots from time
    # Time boundaries mapping:
    # 07:30 AM - 10:30 AM -> Period 1, 2, 3
    if "10:30 am" in c_lower and ("07:30 am" in c_lower or "7:30 am" in c_lower):
        allowed_slots = [1, 2, 3]
    # 10:50 AM - 01:10 PM -> Period 4, 5, 6, 7
    elif "10:50 am" in c_lower and "01:10 pm" in c_lower:
        allowed_slots = [4, 5, 6, 7]
    # 07:30 AM - 01:10 PM -> Period 1-7
    elif "01:10 pm" in c_lower and ("07:30 am" in c_lower or "7:30 am" in c_lower):
        allowed_slots = list(range(1, slots_count + 1))

    # 2. Parse Day Ranges
    if "monday - saturday" in c_lower or "mon - sat" in c_lower:
        allowed_days = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]
    elif "monday - friday" in c_lower or "mon - fri" in c_lower:
        allowed_days = ["Mon", "Tue", "Wed", "Thu", "Fri"]
    elif "monday - wednesday" in c_lower or "mon - wed" in c_lower:
        allowed_days = ["Mon", "Tue", "Wed"]
    elif "wednesday - saturday" in c_lower or "wed - sat" in c_lower:
        allowed_days = ["Wed", "Thu", "Fri", "Sat"]

    # 3. Parse Max Daily Lecture Limits
    max_match = re.search(r'(?:maximum|maximim|max)\s+(\d+)', c_lower)
    if max_match:
        max_per_day = int(max_match.group(1))

    return {
        "allowed_days": allowed_days,
        "allowed_slots": allowed_slots,
        "max_per_day": max_per_day
    }

def parse_user_excel(file_bytes: bytes) -> Dict[str, Any]:
    """
    Parses the new user template: 'Time Table Data.xlsx'
    Required Sheets:
      - 'Teachers Data'
      - 'Sample Class Time Table'
      - 'Lecture Count' (optional, used for capacity checking)
    """
    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    except Exception as e:
        logger.error("excel_load_failed", error=str(e))
        raise IngestionError("Invalid Excel file format. Could not open workbook.")

    # 1. Parse Sample Class Time Table to determine slot size and working days
    if 'Sample Class Time Table' not in wb.sheetnames:
        raise IngestionError("Missing sheet: 'Sample Class Time Table'")
    
    ws_time = wb['Sample Class Time Table']
    max_period = 0
    days_seen = {"Monday", "Tuesday", "Wednesday", "Thursday", "Friday"} # standard defaults

    # Scan periods in column 1
    for r in range(2, ws_time.max_row + 1):
        period_val = ws_time.cell(row=r, column=1).value
        p_num = clean_int(period_val)
        if p_num > max_period:
            max_period = p_num

        # Check for Saturday header on Saturday section
        val_c3 = clean_str(ws_time.cell(row=r, column=3).value)
        if "saturday" in val_c3.lower():
            days_seen.add("Saturday")

    if max_period <= 0:
        max_period = 7 # fallback to 7 periods

    days_list = []
    # Order: Mon-Sat
    for d in ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]:
        if d in days_seen:
            days_list.append(d[:3]) # "Mon", "Tue"...

    # 2. Parse Teachers Data
    if 'Teachers Data' not in wb.sheetnames:
        raise IngestionError("Missing sheet: 'Teachers Data'")
        
    ws_teachers = wb['Teachers Data']
    
    resources_data = []
    availability_data = []
    exceptions_data = [] # empty default
    capacity_data = {"resources": {}, "groups": {}}

    last_teacher = ""
    last_subject = ""
    last_home_group = None
    last_target_class = ""

    teacher_total_lectures = {}

    for r in range(2, ws_teachers.max_row + 1):
        t_name = clean_str(ws_teachers.cell(row=r, column=2).value)
        t_subject = clean_str(ws_teachers.cell(row=r, column=3).value)
        ct_class = clean_str(ws_teachers.cell(row=r, column=4).value)
        target_class = clean_str(ws_teachers.cell(row=r, column=5).value)
        weekly_count_val = ws_teachers.cell(row=r, column=6).value
        total_lectures_val = ws_teachers.cell(row=r, column=7).value
        constraint_str = clean_str(ws_teachers.cell(row=r, column=8).value)

        # Skip empty rows
        if not t_name and not t_subject and not target_class and weekly_count_val is None:
            continue

        if t_name:
            last_teacher = t_name
            last_home_group = ct_class if ct_class else None
            # Store total lectures expected
            if total_lectures_val is not None:
                teacher_total_lectures[last_teacher] = clean_int(total_lectures_val)

        if t_subject:
            last_subject = t_subject

        if target_class:
            last_target_class = target_class
        else:
            target_class = last_target_class

        weekly_count = clean_int(weekly_count_val)
        
        row_id = f"Teachers Data (Row {r})"

        if target_class and weekly_count > 0:
            resources_data.append({
                "resource_name": last_teacher,
                "task": last_subject,
                "home_group": last_home_group,
                "group_name": target_class,
                "weekly_count": weekly_count,
                "row_id": row_id
            })

            # Check if there are constraints specified
            if constraint_str:
                parsed_avail = parse_teacher_constraint(constraint_str, max_period, days_list)
                availability_data.append({
                    "resource_name": last_teacher,
                    "allowed_days": ",".join(parsed_avail["allowed_days"]),
                    "allowed_slots": ",".join(str(s) for s in parsed_avail["allowed_slots"]),
                    "max_per_day": parsed_avail["max_per_day"],
                    "row_id": row_id
                })

    # Save expected hours for validation checks
    capacity_data["resources"] = teacher_total_lectures

    # 3. Parse Lecture Count (optional capacity checks)
    if 'Lecture Count' in wb.sheetnames:
        ws_cap = wb['Lecture Count']
        for r in range(2, ws_cap.max_row + 1):
            class_name = clean_str(ws_cap.cell(row=r, column=1).value)
            total_lectures = ws_cap.cell(row=r, column=2).value
            if class_name and total_lectures is not None:
                capacity_data["groups"][class_name] = clean_int(total_lectures)

    # Make default exceptions: consecutive subject double slots are allowed for EVS and Science, etc.
    # We can default allow consecutive to False, unless specified. Let's leave task exceptions list empty.

    return {
        "config": {
            "Day Names": ",".join(days_list),
            "Slots Per Day": max_period
        },
        "resources_data": resources_data,
        "availability_data": availability_data,
        "exceptions_data": exceptions_data,
        "capacity_data": capacity_data
    }
