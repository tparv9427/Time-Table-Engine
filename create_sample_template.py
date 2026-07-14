import openpyxl

def create_template():
    wb = openpyxl.Workbook()

    # 1. Sheet: Organization Config
    ws_cfg = wb.active
    ws_cfg.title = "Organization Config"
    ws_cfg.append(["Key", "Value"])
    ws_cfg.append(["Day Names", "Mon,Tue,Wed,Thu,Fri"])
    ws_cfg.append(["Slots Per Day", "6"])
    ws_cfg.append(["Slot Times", "9:00 AM,10:00 AM,11:00 AM,12:00 PM,1:00 PM,2:00 PM"])

    # 2. Sheet: Resources Data
    ws_res = wb.create_sheet(title="Resources Data")
    ws_res.append(["Resource Name", "Task/Subject", "Home Group", "Group", "Weekly Count"])
    # 15 allocations for Class 10A, 15 allocations for Class 10B
    ws_res.append(["Teacher A", "Math", "Class 10A", "Class 10A", 5])
    ws_res.append(["Teacher B", "Science", "", "Class 10A", 5])
    ws_res.append(["Teacher C", "English", "", "Class 10A", 5])
    ws_res.append(["Teacher D", "Math", "Class 10B", "Class 10B", 5])
    ws_res.append(["Teacher E", "Science", "", "Class 10B", 5])
    ws_res.append(["Teacher F", "English", "", "Class 10B", 5])

    # 3. Sheet: Resource Availability
    ws_avail = wb.create_sheet(title="Resource Availability")
    ws_avail.append(["Resource Name", "Allowed Days", "Allowed Slots", "Max Per Day"])
    ws_avail.append(["Teacher A", "Mon,Tue,Wed", "1,2,3,4,5,6", 2])
    ws_avail.append(["Teacher B", "Tue,Wed,Thu", "1,2,3,4,5,6", 3])
    ws_avail.append(["Teacher C", "Wed,Thu,Fri", "1,2,3,4,5,6", 2])
    ws_avail.append(["Teacher D", "Mon,Tue,Wed", "1,2,3,4,5,6", 3])
    ws_avail.append(["Teacher E", "Tue,Wed,Thu", "1,2,3,4,5,6", 2])
    ws_avail.append(["Teacher F", "Wed,Thu,Fri", "1,2,3,4,5,6", 2])

    # 4. Sheet: Task Exceptions
    ws_except = wb.create_sheet(title="Task Exceptions")
    ws_except.append(["Task/Subject", "Applicable Groups", "Allow Consecutive"])
    ws_except.append(["Math", "Class 10A,Class 10B", "N"])
    ws_except.append(["Science", "*", "Y"])

    # 5. Sheet: Capacity Reconciliation
    ws_reconcile = wb.create_sheet(title="Capacity Reconciliation")
    ws_reconcile.append(["Category", "Name", "Expected Weekly Hours"])
    ws_reconcile.append(["Resource", "Teacher A", 5])
    ws_reconcile.append(["Resource", "Teacher B", 5])
    ws_reconcile.append(["Resource", "Teacher C", 5])
    ws_reconcile.append(["Resource", "Teacher D", 5])
    ws_reconcile.append(["Resource", "Teacher E", 5])
    ws_reconcile.append(["Resource", "Teacher F", 5])
    ws_reconcile.append(["Group", "Class 10A", 15])
    ws_reconcile.append(["Group", "Class 10B", 15])

    wb.save("sample_timetable_template.xlsx")
    print("Created sample_timetable_template.xlsx successfully!")

if __name__ == "__main__":
    create_template()
